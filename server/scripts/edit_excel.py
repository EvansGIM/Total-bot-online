#!/usr/bin/env python3
"""
Excel 파일 직접 편집 스크립트 (서식 보존)
openpyxl을 사용하여 기존 파일의 셀 값만 수정
"""

import sys
import json
from openpyxl import load_workbook


def edit_excel(file_path, cell_updates):
    """
    Excel 파일의 특정 셀만 수정 (서식 유지)

    Args:
        file_path: Excel 파일 경로
        cell_updates: [{"sheet": 시트명/인덱스, "row": 행, "col": 열, "value": 값}, ...]

    Returns:
        dict: {"success": True/False, "updated": 수정된 셀 수, "error": 에러메시지}
    """
    try:
        # 파일 열기 (서식 유지)
        wb = load_workbook(file_path)
        updated_count = 0

        for update in cell_updates:
            # 시트 선택 (이름 또는 인덱스)
            sheet_ref = update.get('sheet')
            if sheet_ref is None:
                # 기본: 2번째 시트 (인덱스 1)
                ws = wb.worksheets[1] if len(wb.worksheets) > 1 else wb.worksheets[0]
            elif isinstance(sheet_ref, int):
                ws = wb.worksheets[sheet_ref]
            else:
                ws = wb[sheet_ref]

            row = update['row']
            col = update['col']
            value = update['value']

            # 셀 값만 수정 (서식은 그대로 유지됨)
            ws.cell(row=row, column=col, value=value)
            updated_count += 1

        # 같은 파일에 저장 (덮어쓰기)
        wb.save(file_path)
        wb.close()

        return {"success": True, "updated": updated_count}

    except Exception as e:
        return {"success": False, "error": str(e)}


if __name__ == '__main__':
    if len(sys.argv) != 3:
        print(json.dumps({
            "success": False,
            "error": "Usage: edit_excel.py <file_path> <cell_updates_json_file>"
        }))
        sys.exit(1)

    file_path = sys.argv[1]
    cell_updates_file = sys.argv[2]  # JSON 파일 경로

    # JSON 파일에서 데이터 로드
    try:
        with open(cell_updates_file, 'r', encoding='utf-8') as f:
            cell_updates = json.load(f)
    except FileNotFoundError:
        print(json.dumps({
            "success": False,
            "error": f"JSON 파일을 찾을 수 없음: {cell_updates_file}"
        }))
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(json.dumps({
            "success": False,
            "error": f"JSON 파싱 오류: {str(e)}"
        }))
        sys.exit(1)

    result = edit_excel(file_path, cell_updates)
    print(json.dumps(result, ensure_ascii=False))

    if not result['success']:
        sys.exit(1)

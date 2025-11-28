#!/usr/bin/env python3
"""
견적서 Excel 파일 자동 작성 스크립트
"""

import sys
import openpyxl
import json

def fill_excel(file_path, category, search_tags, size, weight):
    """
    Excel 파일에 데이터 자동 채우기
    
    Args:
        file_path: Excel 파일 경로
        category: 선택한 카테고리
        search_tags: 검색태그 문자열
        size: 사이즈 문자열 (예: "10x20x30")
        weight: 무게 문자열 (예: "500g")
    """
    print(f"📝 파일 열기: {file_path}")
    
    # Excel 파일 로드
    wb = openpyxl.load_workbook(file_path)
    
    # 2번째 시트 (index 1)
    ws = wb.worksheets[1]
    sheet_name = wb.sheetnames[1]
    print(f"📊 시트명: {sheet_name}")
    
    # 5행에서 헤더 찾기
    headers = {}
    for col in range(1, 100):
        cell = ws.cell(row=5, column=col)
        if cell.value:
            headers[cell.value] = col
    
    print(f"📋 헤더 {len(headers)}개 발견")
    
    # 9행에 데이터 작성
    data_row = 9
    
    # B9에 카테고리 입력
    ws.cell(row=data_row, column=2, value=category)
    print(f"✅ B9에 카테고리 입력: {category}")
    
    # 검색태그
    if '검색태그' in headers:
        col = headers['검색태그']
        ws.cell(row=data_row, column=col, value=search_tags)
        print(f"✅ 검색태그 입력 (col {col}): {search_tags}")
    
    # 사이즈
    if '한 개 단품 포장 사이즈' in headers:
        col = headers['한 개 단품 포장 사이즈']
        ws.cell(row=data_row, column=col, value=size)
        print(f"✅ 사이즈 입력 (col {col}): {size}")
    
    # 무게
    if '한 개 단품 포장 무게' in headers:
        col = headers['한 개 단품 포장 무게']
        ws.cell(row=data_row, column=col, value=weight)
        print(f"✅ 무게 입력 (col {col}): {weight}")
    
    # 저장
    wb.save(file_path)
    print(f"💾 파일 저장 완료: {file_path}")

if __name__ == '__main__':
    if len(sys.argv) != 6:
        print("Usage: fill_excel.py <file_path> <category> <search_tags> <size> <weight>")
        sys.exit(1)
    
    file_path = sys.argv[1]
    category = sys.argv[2]
    search_tags = sys.argv[3]
    size = sys.argv[4]
    weight = sys.argv[5]
    
    fill_excel(file_path, category, search_tags, size, weight)

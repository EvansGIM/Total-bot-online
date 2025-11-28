#!/usr/bin/env python3
"""
견적서 Excel 파일 자동 작성 스크립트 (간단 버전)
선택한 상품들의 옵션을 9행부터 추가
"""

import sys
import openpyxl
import json

def fill_quotation_excel(excel_path, products_data, category, search_tags, size, weight):
    """
    Excel 파일에 상품 데이터 자동 채우기

    Args:
        excel_path: Excel 파일 경로
        products_data: 선택한 상품들의 데이터 (JSON 배열)
        category: 선택한 카테고리
        search_tags: 검색태그 문자열
        size: 사이즈 문자열 (예: "10x20x30")
        weight: 무게 문자열 (예: "500g")
    """
    print(f"📝 파일 열기: {excel_path}")

    # Excel 파일 로드
    wb = openpyxl.load_workbook(excel_path)

    # 2번째 시트 (index 1)
    ws = wb.worksheets[1]
    sheet_name = wb.sheetnames[1]
    print(f"📊 시트명: {sheet_name}")

    # 5행에서 헤더 찾기
    headers = {}
    for col in range(1, 100):
        cell = ws.cell(row=5, column=col)
        if cell.value:
            header_name = str(cell.value).strip().replace('\n', ' ')
            headers[header_name] = col

    print(f"📋 헤더 {len(headers)}개 발견")
    print(f"   발견된 헤더: {list(headers.keys())[:10]}")  # 처음 10개만 출력

    # 헤더명 매핑 (유사한 이름 찾기)
    def find_header(keywords):
        """여러 키워드 중 하나라도 헤더에 포함되면 반환"""
        for keyword in keywords:
            for header_name, col in headers.items():
                if keyword in header_name:
                    return col
        return None

    # 주요 헤더 찾기
    col_category = find_header(['카테고리'])
    col_product_name = find_header(['상품명'])
    col_option1 = find_header(['옵션1', '색상', '패션의류'])
    col_option2 = find_header(['옵션2', '사이즈'])
    col_search_tag = find_header(['검색태그'])
    col_weight = find_header(['무게'])
    col_size = find_header(['사이즈', '포장 사이즈'])
    col_price = find_header(['공급가', '판매가'])

    print(f"\n📍 매핑된 열:")
    print(f"   카테고리: {col_category}")
    print(f"   상품명: {col_product_name}")
    print(f"   옵션1: {col_option1}")
    print(f"   옵션2: {col_option2}")
    print(f"   검색태그: {col_search_tag}")
    print(f"   무게: {col_weight}")
    print(f"   사이즈: {col_size}")
    print(f"   가격: {col_price}")

    # 9행부터 데이터 작성
    current_row = 9

    # 선택한 상품들의 옵션을 순회
    for product in products_data:
        product_title = product.get('title') or product.get('titleCn') or '제목 없음'
        options = product.get('results', [])

        if not options:
            # 옵션이 없는 경우 상품 1개만 추가
            if col_category:
                ws.cell(row=current_row, column=col_category, value=category)
            if col_product_name:
                ws.cell(row=current_row, column=col_product_name, value=product_title)
            if col_search_tag:
                ws.cell(row=current_row, column=col_search_tag, value=search_tags)
            if col_weight:
                ws.cell(row=current_row, column=col_weight, value=weight)
            if col_size:
                ws.cell(row=current_row, column=col_size, value=size)
            if col_price:
                price = product.get('salePrice') or product.get('basePrice') or 0
                ws.cell(row=current_row, column=col_price, value=price)

            print(f"✅ Row {current_row}: {product_title} (옵션 없음)")
            current_row += 1
        else:
            # 각 옵션마다 한 줄씩 추가
            for option in options:
                if col_category:
                    ws.cell(row=current_row, column=col_category, value=category)
                # 옵션값을 먼저 가져옴
                opt1 = option.get('optionName1') or option.get('optionName1Cn') or ''
                opt2 = option.get('optionName2') or option.get('optionName2Cn') or ''

                # 상품명 = 제품명 + 옵션1 + 옵션2 (스페이스 구분)
                if col_product_name:
                    combined_name = ' '.join(filter(None, [product_title, opt1, opt2]))
                    ws.cell(row=current_row, column=col_product_name, value=combined_name)
                if col_option1:
                    ws.cell(row=current_row, column=col_option1, value=opt1)
                if col_option2:
                    ws.cell(row=current_row, column=col_option2, value=opt2)
                if col_search_tag:
                    ws.cell(row=current_row, column=col_search_tag, value=search_tags)
                if col_weight:
                    ws.cell(row=current_row, column=col_weight, value=weight)
                if col_size:
                    ws.cell(row=current_row, column=col_size, value=size)
                if col_price:
                    price = option.get('price', 0)
                    ws.cell(row=current_row, column=col_price, value=price)

                opt1_display = option.get('optionName1') or option.get('optionName1Cn') or '-'
                opt2_display = option.get('optionName2') or option.get('optionName2Cn') or '-'
                print(f"✅ Row {current_row}: {product_title} | {opt1_display} | {opt2_display}")
                current_row += 1

    # 저장
    wb.save(excel_path)
    print(f"\n💾 파일 저장 완료: {excel_path}")
    print(f"📊 총 {current_row - 9}개 옵션 추가됨")

if __name__ == '__main__':
    if len(sys.argv) != 7:
        print("Usage: fill_quotation.py <excel_path> <products_json> <category> <search_tags> <size> <weight>")
        sys.exit(1)

    excel_path = sys.argv[1]
    products_json = sys.argv[2]  # JSON 문자열
    category = sys.argv[3]
    search_tags = sys.argv[4]
    size = sys.argv[5]
    weight = sys.argv[6]

    # JSON 파싱
    try:
        products_data = json.loads(products_json)
    except json.JSONDecodeError as e:
        print(f"❌ JSON 파싱 오류: {e}")
        sys.exit(1)

    fill_quotation_excel(excel_path, products_data, category, search_tags, size, weight)
